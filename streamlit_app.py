import streamlit as st
import pandas as pd
import json
from io import BytesIO
import io
from mistralai import Mistral
import base64
from openpyxl import Workbook
import numpy as np
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font,PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, get_column_letter
from google import genai

tab1,  tab5 = st.tabs(["AI PDF Extractor", "Excel Ledger Decoder"])

# 🔐 Set your API key

with tab1:
    #st.set_page_config(page_title="📄 Mistral PDF Extractor", layout="wide")
    st.title("📄 Upload PDFs for AI-Powered Bank Statement Extraction")
    
    st.title("🔐 API Key Input")
    
    # Input box for API key
    api_key = st.text_input("Enter your API key:", type="password")
    client = Mistral(api_key = api_key)
    
    model = "mistral-small-latest"
    
    
    st.sidebar.header("🧠 Prompt Settings")
    prompt_type = st.sidebar.selectbox(
        "Choose prompt style:",
        ["Long Prompt", "Short Prompt", "Custom Prompt"]
    )
    
    custom_prompt = ""
    run_custom = False
    if prompt_type == "Custom Prompt":
        custom_prompt = st.sidebar.text_area("Enter your custom prompt")
        run_custom = st.sidebar.button("🚀 Run Custom Prompt")
    
    if prompt_type != "Custom Prompt" or run_custom:
        if prompt_type == "Long Prompt":
            prompt_text = (
                "You are a meticulous data extraction specialist with extensive experience in processing financial documents. Your expertise lies in accurately extracting and organizing transaction data from bank statements while ensuring the integrity of the information. Your task is to extract and structure transaction data from a bank statement. The output in a json format: [{\"Company Name\": \"Company Name should be same company for the same statement\",\"Bank Account Number\": \"Bank Account Number\",\"Currency\": \"Currency\",\"Date\": \"Date (YYYY-MM-DD format\",\"Description\": \"Description\"\"Deposit\": \"Deposit / \"\"Withdrawal\": \"Withdrawal\"\"Balance\": \"Balance\"}]. Ensure that the Deposit and Withdrawal categories from the bank statement are accurately represented in the table. It is crucial to determine the column order by locating the headers for 'Deposit' and 'Withdrawal.' Trust the positions of the columns strictly for credit vs. debit classification, not the description text. Capture the “Statement Balance” after processing the last transaction of each date, and verify it against your computed running balance. If there are discrepancies, re-evaluate your data capture and make necessary adjustments. Begin by obtaining the brought forward balance, which may sometimes be stated in the first line of the balance section. This balance should not be categorized as a deposit but recorded as balance in the brought forward line. For transactions with multiple entries on a single date, ensure that each transaction is captured as a separate row. Calculate a running balance for each transaction and validate the mathematical progression: Previous Balance ± Amount = New Balance. The new balance must match the bank statement; if inconsistencies arise, flag them for review or correct the classification. Proceed with a detailed, manual, line-by-line extraction. The concluding balance should be recalculated after each date's last transaction to confirm its accuracy. If the ending balance does not match the bank statement, recheck the document for any missing or incorrect data. Please adhere to the following constraints during the extraction process:  - Do not provide sample transactions or summaries. - Ensure no transactions are skipped due to volume or complexity. - Verify the total transaction count, investigate for date gaps, confirm the final balance matches the statement closing balance, and ensure all pages have been processed. Your output must be comprehensive and accurate, reflecting every single transaction without exceptions."
            )
        elif prompt_type == "Short Prompt":
            prompt_text = (
                "Your task is to extract and structure transaction data from a bank statement. The output in a json format: [{\"Company Name\": \"Company Name should be same company for the same statement\",\"Bank Account Number\": \"Bank Account Number\",\"Currency\": \"Currency\",\"Date\": \"Date (YYYY-MM-DD format\",\"Description\": \"Description\"\"Deposit\": \"Deposit / \"\"Withdrawal\": \"Withdrawal\"\"Balance\": \"Balance or Balance in original currrency\"}]. Ignore account summary"
            )
        else:        
            prompt_text = custom_prompt + "The output in a json format: [{\"Company Name\": \"Company Name should be same company for the same statement\",\"Bank Account Number\": \"Bank Account Number\",\"Currency\": \"Currency\",\"Date\": \"Date (YYYY-MM-DD format\",\"Description\": \"Description\"\"Deposit\": \"Deposit / \"\"Withdrawal\": \"Withdrawal\"\"Balance\": \"Balance or Balance in original currrency\"}]"
    
    def display_pdf_inline_old(file):
        base64_pdf = base64.b64encode(file.read()).decode("utf-8")
        pdf_viewer = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="800px" type="application/pdf"></iframe>'
        st.markdown(pdf_viewer, unsafe_allow_html=True)
    
    def display_pdf_inline(file):
        base64_pdf = base64.b64encode(file.read()).decode("utf-8")
        pdf_viewer = f'''
            <embed src="data:application/pdf;base64,{base64_pdf}" width="100%" height="800px" type="application/pdf">
        '''
        st.components.v1.html(pdf_viewer, height=800)
        
    uploaded_files = st.file_uploader("Upload PDF files", type="pdf", accept_multiple_files=True)
    
    if uploaded_files:
        all_data = pd.DataFrame()
        for file in uploaded_files:
            st.write(f"📤 Uploading: {file.name}")
            file_bytes = file.read()
    
            # Upload to Mistral for OCR
            uploaded_pdf = client.files.upload(
                file={"file_name": file.name, "content": file_bytes},
                purpose="ocr"
            )
    
            # Get signed URL for document
            signed_url = client.files.get_signed_url(file_id=uploaded_pdf.id)
    
            # 🧾 Prompt for structured extraction
            messages = [{
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt_text
                    },
                    {
                        "type": "document_url",
                        "document_url": signed_url.url
                    }
                ]
            }]
    
            chat_response = client.chat.complete(model=model, messages=messages)
            content = chat_response.choices[0].message.content
    
            # Extract JSON from response
            try:
                start = content.find("[")
                end = content.rfind("]") + 1
                json_data = content[start:end]
                records = json.loads(json_data)
    
                for record in records:
                    record["Pdf Name"] = file.name
    
                df = pd.DataFrame(records)
                all_data = pd.concat([all_data, df], ignore_index=True)
                st.success(f"✅ Extracted {len(df)} records from {file.name}")
    
            except Exception as e:
                st.error(f"❌ Failed to parse response from {file.name}: {e}")
    
        # Display and download
            # 🔲 Side-by-side layout
            col1, col2 = st.columns([1, 1])
            with col1:
                    st.markdown("#### 📄 Raw PDF")
                    file.seek(0)  # Reset pointer for display
                    display_pdf_inline(file)
    
            with col2:
                    st.markdown("#### 🧠 Extracted Data")
        #            st.table(df)
                    st.dataframe(df)
        st.title("Full data:")
        st.dataframe(all_data)
        output = BytesIO()
        all_data.to_excel(output, index=False)
        st.download_button("📥 Download Extracted Data", output.getvalue(), file_name="mistral_extracted.xlsx")

with tab5:

    st.title("Excel Ledger Decoder")
    st.markdown("""
    Upload your accounting Excel file (where amounts are spread across columns H, I, J...)  
    → The app will **decode** it into clean rows: one transaction line per account.
    """)

    # File uploader
    uploaded_file = st.file_uploader("Choose your Excel file", type=["xlsx", "xls", "xlsm"])

    if uploaded_file is not None:
        try:
            # Read all sheets (in case there are multiple)
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names

            # Let user choose sheet
            selected_sheet = st.selectbox("Select sheet", sheet_names)

            # Read the selected sheet
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=7)
            # Copy value from A10 to B10
            df.iloc[1, 1] = df.iloc[1, 0]   # row 10 → index 9; col B → index 1; col A → index 0
            df = df.drop(index=0)
            df = df.drop(columns=["Check"])
            # Reset index if needed
            df = df.reset_index(drop=True)
            st.success(f"Loaded sheet **{selected_sheet}** with {len(df)} rows and {len(df.columns)} columns.")
            df.loc[df["Date"] == "Grand Total", "Description"] = "Grand Total"
            # Auto-detect where the fixed columns end (usually column G = index 6)
            # We'll assume everything from column H onward contains account names in headers and amounts in cells
            fixed_cols_count = 7  # A to G → 7 columns

            if len(df.columns) < fixed_cols_count:
                st.error("File has fewer than 7 columns. Is this the correct file?")
            else:
                # Show preview of original data
                st.subheader("Preview of Original Data")
                st.dataframe(df.head(10))

                # Decode logic
                new_rows = []
                fixed_column_names = df.columns[:fixed_cols_count]

                for _, row in df.iterrows():
                    fixed_part = row.iloc[:fixed_cols_count].tolist()

                    # From column H onward
                    for col_idx in range(fixed_cols_count, len(row)):
                        account_name = df.columns[col_idx]  # Header = account name
                        amount = row.iloc[col_idx]

                        if pd.notna(amount) and str(amount).strip() != "":
                            amount = pd.to_numeric(amount, errors='coerce')
                            if amount is not None:  # safety
                                new_rows.append(fixed_part + [account_name.strip(), amount])

                # Create result DataFrame
                result_df = pd.DataFrame(
                    new_rows,
                    columns=list(fixed_column_names) + ["Account Type", "Value"]
                )

                st.subheader(f"Decoded Result – {len(result_df)} transaction lines")
                st.dataframe(result_df)

                # Offer download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Decoded')
                output.seek(0)

                base_name = os.path.splitext(uploaded_file.name)[0]
                
                # Force .xlsx extension
                file_name = f"DECODED_{base_name}.xlsx"
                
                st.download_button(
                    label="Download Decoded Excel File",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.balloons()

        except Exception as e:
            st.error(f"Error: {e}")
            st.write("Make sure your file is a valid Excel (.xlsx) and try again.")
